
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import  Compra, DetalleCompra, Producto, Proveedor
from .forms import  CompraForm, DetalleCompraForm
from openpyxl.drawing.image import Image
from reportlab.lib.utils import ImageReader
from django.views.decorators.cache import never_cache
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from django.forms import inlineformset_factory
from django.http import JsonResponse
# Create your views here.
@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')

@login_required
@never_cache
def crear_compra(request):
    DetalleCompraFormSet = inlineformset_factory(
        Compra, 
        DetalleCompra, 
        form=DetalleCompraForm, 
        extra=1,  # Puedes cambiar este número si no quieres formularios adicionales por defecto
        can_delete=True
    )
    
    if request.method == 'POST':
        form = CompraForm(request.POST)
        formset = DetalleCompraFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            compra = form.save()
            formset.instance = compra # Asociar la venta a los detalles
            formset.save()

            # Imprimir el número de formularios en el formset
            print("Número de formularios en el formset:", formset.total_form_count())
            
            # Calcular y guardar el valor total
            valor_total = sum(detalle.subtotal for detalle in compra.detalles.all())
            compra.valor_total = valor_total
            compra.save()
            
            return redirect('gestionar_compras')
        else:
            # Depuración de errores
            print(form.errors, formset.errors)  
    else:
        form = CompraForm()
        formset = DetalleCompraFormSet()
    
    return render(request, 'crear_compra.html', {'form': form, 'formset': formset})


@login_required
@never_cache
def gestionar_compras(request):
    compras = Compra.objects.all()
    return render(request, 'gestionar_compras.html', {'compras': compras})

@login_required
@never_cache
def detalle_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    return render(request, 'detalle_compra.html', {'Compra': compra})

@login_required
@never_cache
def editar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    DetalleCompraFormSet = inlineformset_factory(Compra, DetalleCompra, form=DetalleCompraForm, extra=1)
    
    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit=False)
            formset = DetalleCompraFormSet(request.POST, instance=compra)
            if formset.is_valid():
                compra.save()
                formset.save()
                # Recalcular y guardar el valor total
                valor_total = sum(detalle.subtotal for detalle in compra.detalles.all())
                compra.valor_total = valor_total
                compra.save()
                return redirect('gestionar_compras')
    else:
        form = CompraForm(instance=compra)
        formset = DetalleCompraFormSet(instance=compra)
    
    return render(request, 'editar_compra.html', {'form': form, 'formset': formset})

@login_required
@never_cache
def eliminar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    if request.method == 'POST':
        compra.delete()
        messages.success(request, 'La compra ha sido eliminada exitosamente.')
        return redirect('gestionar_compras')
    return render(request, 'eliminar_compra.html', {'compra': compra})




def obtener_proveedor(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
        proveedor = producto.proveedor  # Asegúrate de que 'proveedor' es el campo correcto
        return JsonResponse({'proveedor': proveedor})
    except Producto.DoesNotExist:
        return JsonResponse({'proveedor': ''}, status=404)




@never_cache
def obtener_precio_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    return JsonResponse({'precio': producto.precio})
def reporte_compras_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.5)  # Ajusta la transparencia (0.1 = 10% opaco)
            img = ImageReader(watermark_path)
            iw, ih = img.getSize()
            aspect = ih / float(iw)
            p.drawImage(img, x=0, y=0, width=width, height=height*aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "SAMSAMANA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Reporte de Compras")
    y_position -= 50

    # Crear tabla con datos de compras
    column_widths = [30, 80, 80, 80, 80, 60]
    headers = ["ID", "Fecha Compra", "Total Compra", "Cantidad Producto", "Proveedor", "Estado"]
    data = [headers]

    compras = Compra.objects.all()
    for compra in compras:
        data.append([
            str(compra.id),
            compra.fecha_Compra.strftime('%Y-%m-%d %H:%M:%S'),
            str(compra.total_Compra),
            str(compra.cantidad_Producto),
            compra.proveedor_Id.nombre,  # Asegúrate de que `nombre` es un atributo de `Proveedor`
            'Activo' if compra.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (-1, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])
    table.setStyle(style)
    
    # Ajustar la posición de la tabla
    table_width = sum(column_widths)
    table_x = (width - table_width - 2 * margin) / 2 + margin
    table_y = y_position - len(data) * row_height
    table.wrapOn(p, table_width, height - 2 * margin)
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras.pdf"'
    return response



def reporte_compras_excel(request):
    # Crear un archivo Excel en memoria
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Compras"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Añadir logo (más pequeño)
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 80  # Ajustar el tamaño del logo
        img.height = 40
        ws.add_image(img, 'A1')

    # Añadir título
    ws.merge_cells('B1:H1')
    title_cell = ws['B1']
    title_cell.value = "TABLA COMPRAS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Añadir encabezados
    headers = ["ID", "Fecha Compra", "Total Compra", "Cantidad Producto", "Proveedor", "Estado"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Añadir datos de compras
    compras = Compra.objects.all()
    for row_num, compra in enumerate(compras, 4):
        ws.cell(row=row_num, column=1, value=compra.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=compra.fecha_Compra.strftime('%Y-%m-%d %H:%M:%S')).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=compra.total_Compra).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=compra.cantidad_Producto).alignment = alignment_center
        ws.cell(row=row_num, column=5, value=compra.proveedor_Id.nombre).alignment = alignment_center  # Verificar que 'nombre' sea el atributo correcto
        ws.cell(row=row_num, column=6, value='Activo' if compra.estado else 'Inactivo').alignment = alignment_center

    # Ajustar el ancho de las columnas
    column_widths = [5, 20, 15, 20, 30, 10]  # Ajustar los anchos de columna según sea necesario
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Guardar el archivo en el buffer
    wb.save(buffer)
    buffer.seek(0)

    # Preparar la respuesta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras_samsamana.xlsx"'
    return response
